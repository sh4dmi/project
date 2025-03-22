#!/usr/bin/env python3
"""
Excel Functions Module
=====================

This module provides a comprehensive set of functions for Excel operations
including reading, writing, and clearing data, as well as reward calculation
for operation success/failure.

The ExcelHandler class encapsulates all Excel operations and provides
both direct API calls and JSON-based operations with reward feedback.
"""

import os
import json
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('excel_functions')

class ExcelHandler:
    """
    ExcelHandler provides a comprehensive set of functions for Excel operations
    including reading, writing, clearing data, and processing JSON commands.
    
    Attributes:
        filename (str): The name of the Excel file to work with.
        workbook (Workbook): The openpyxl Workbook object.
        sheet (Worksheet): The active worksheet in the workbook.
    """
    
    def __init__(self, filename="excel_operations.xlsx"):
        """
        Initialize the ExcelHandler with a specified filename.
        
        Args:
            filename (str): The name of the Excel file to work with.
        """
        self.filename = filename
        logger.info(f"Initializing ExcelHandler with file: {filename}")
        
        # Create a new workbook or load existing one
        if os.path.exists(filename):
            try:
                self.workbook = load_workbook(filename)
                logger.info(f"Loaded existing workbook: {filename}")
            except Exception as e:
                logger.error(f"Error loading workbook: {str(e)}")
                self.workbook = Workbook()
                logger.info("Created new workbook due to loading error")
        else:
            self.workbook = Workbook()
            logger.info("Created new workbook")
        
        # Use the active sheet
        self.sheet = self.workbook.active
        
        # Save the workbook
        self.workbook.save(filename)
    
    #
    # HELPER METHODS
    #
    
    def _get_col_index(self, col_index):
        """
        Convert column index to numerical index if it's a letter.
        
        Args:
            col_index: Column index (can be int, str number, or column letter like 'A')
            
        Returns:
            int: Numerical column index or None if invalid
        """
        try:
            if isinstance(col_index, int):
                if col_index <= 0:
                    logger.error(f"Column index must be positive, got {col_index}")
                    return None
                logger.debug(f"Using numeric column index: {col_index}")
                return col_index
            elif isinstance(col_index, str):
                if col_index.isdigit():
                    numeric_index = int(col_index)
                    if numeric_index <= 0:
                        logger.error(f"Column index must be positive, got {numeric_index}")
                        return None
                    logger.debug(f"Converted string digit '{col_index}' to numeric column index: {numeric_index}")
                    return numeric_index
                else:
                    # Handle column letters (A, B, AA, etc.)
                    try:
                        numeric_index = column_index_from_string(col_index.upper())
                        logger.debug(f"Converted column letter '{col_index}' to numeric index: {numeric_index}")
                        return numeric_index
                    except Exception as e:
                        logger.error(f"Invalid column letter: '{col_index}', error: {str(e)}")
                        return None
            else:
                logger.error(f"Invalid column index type: {type(col_index).__name__}")
                return None
        except Exception as e:
            logger.error(f"Error in _get_col_index: {str(e)}")
            return None
    
    def _validate_row_index(self, row_index):
        """
        Validate that a row index is positive or "next_available".
        
        Args:
            row_index: Row index to validate
            
        Returns:
            bool: True if valid, False otherwise
        """
        if row_index == "next_available":
            return True
        
        if isinstance(row_index, str) and row_index.isdigit():
            row_index = int(row_index)
        
        if not isinstance(row_index, int):
            return False
        
        return row_index > 0
    
    def _get_actual_row_index(self, row_index):
        """
        Convert "next_available" to an actual row index.
        
        Args:
            row_index: Row index or "next_available"
            
        Returns:
            int: Actual row index to use
        """
        if row_index == "next_available":
            # Find the next available row (after the last non-empty row)
            max_row = self.sheet.max_row
            return max_row + 1
        
        if isinstance(row_index, str) and row_index.isdigit():
            return int(row_index)
        
        return row_index
    
    def _format_cell_reference(self, row_index, col_index):
        """
        Format a cell reference in a human-readable way for LLM consumption.
        
        Args:
            row_index (int): Row index (1-based)
            col_index (int or str): Column index (1-based) or letter (A, B, etc.)
            
        Returns:
            str: Human-readable cell reference like "cell at row 2, column B (2,B)"
        """
        # Get the column letter if col_index is numeric
        if isinstance(col_index, int):
            col_letter = get_column_letter(col_index)
        else:
            # If it's already a letter, use it directly
            if isinstance(col_index, str) and col_index.isalpha():
                col_letter = col_index.upper()
            else:
                # Try to convert to int and then to letter
                try:
                    num_col = self._get_col_index(col_index)
                    if num_col:
                        col_letter = get_column_letter(num_col)
                    else:
                        col_letter = str(col_index)
                except:
                    col_letter = str(col_index)
        
        # Create a human-readable format that describes the cell location clearly
        return f"cell at row {row_index}, column {col_letter} ({row_index},{col_letter})"
    
    def _validate_parameters(self, params, required_params):
        """
        Validate that the required parameters are present in the params dict.
        
        Args:
            params (dict): Parameters to validate
            required_params (list): List of required parameter names
            
        Returns:
            bool: True if all required parameters are present, False otherwise
        """
        if not params:
            return len(required_params) == 0
            
        for param in required_params:
            if param not in params:
                return False
        return True
    
    #
    # EXCEL OPERATIONS
    #
    
    def clear_sheet(self):
        """
        Clear all data from the active sheet.
        
        Returns:
            tuple: (success, message)
                - success (bool): True if operation succeeded, False otherwise
                - message (str): Success or error message
        """
        try:
            # Get sheet dimensions before clearing
            max_row = self.sheet.max_row
            max_col = self.sheet.max_column
            sheet_name = self.sheet.title
            
            # Create a new worksheet to replace the existing one
            # This is more reliable than deleting rows for a complete reset
            ws_name = self.sheet.title
            self.workbook.remove(self.sheet)
            self.sheet = self.workbook.create_sheet(ws_name)
            self.workbook.active = self.sheet
            
            success_msg = f"Sheet '{sheet_name}' cleared. Removed all data ({max_row} rows by {max_col} columns). A new empty sheet has been created."
            logger.info("Sheet cleared successfully (recreated)")
            
            return True, success_msg
        except Exception as e:
            error_msg = f"Error clearing sheet: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def add_row(self, row_index, text):
        """
        Add a new row at the specified index.
        
        Args:
            row_index (int or str): Index of the row to add, or "next_available"
            text (str): Text to add in the first cell of the row
            
        Returns:
            tuple: (success, message)
                - success (bool): True if operation succeeded, False otherwise
                - message (str): Success or error message
        """
        try:
            # Validate row_index
            if not self._validate_row_index(row_index):
                error_msg = f"Invalid row index: {row_index}. Row index must be positive integer or 'next_available'."
                logger.error(error_msg)
                return False, error_msg
            
            # Get actual row index
            actual_row_index = self._get_actual_row_index(row_index)
            
            # Insert a row and add text to the first cell
            self.sheet.insert_rows(actual_row_index)
            self.sheet.cell(row=actual_row_index, column=1).value = text
            
            # Create a cell reference for the first cell where text was added
            cell_ref = self._format_cell_reference(actual_row_index, 'A')
            
            success_msg = f"New row inserted at position {actual_row_index}. Text '{text}' added to {cell_ref}"
            logger.info(success_msg)
            
            # Save the workbook
            self.workbook.save(self.filename)
            
            return True, success_msg
        except Exception as e:
            error_msg = f"Error adding row: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def write_cell(self, row_index, col_index, text):
        """
        Write text to a specific cell.
        
        Args:
            row_index (int): Row index (1-based)
            col_index (int or str): Column index (1-based) or letter (A, B, etc.)
            text: Content to write to the cell
            
        Returns:
            tuple: (success, message)
                - success (bool): True if operation succeeded, False otherwise
                - message (str): Success or error message
        """
        try:
            # Convert string row_index to int if it's a digit
            if isinstance(row_index, str) and row_index.isdigit():
                row_index = int(row_index)
                logger.info(f"Converted string row_index to int: {row_index}")
            
            # Validate row_index
            if not self._validate_row_index(row_index):
                error_msg = f"Invalid row index: {row_index}. Row index must be positive integer."
                logger.error(error_msg)
                return False, error_msg
            
            # Get the column index if it's a letter
            logger.info(f"Processing column index: {col_index} (type: {type(col_index).__name__})")
            num_col_index = self._get_col_index(col_index)
            if num_col_index is None:
                error_msg = f"Invalid column index: {col_index}"
                logger.error(error_msg)
                return False, error_msg
            
            # Get column letter for better readability
            col_letter = get_column_letter(num_col_index)
            
            # Log explicit target coordinates before writing
            logger.info(f"TARGET CELL: Row={row_index}, Column={num_col_index} (Letter: {col_letter})")
            
            # Create formatted cell reference
            cell_ref = self._format_cell_reference(row_index, col_letter)
            
            # Write to the cell - ONLY to the specified cell, nothing else
            target_cell = self.sheet.cell(row=row_index, column=num_col_index)
            old_value = target_cell.value
            target_cell.value = text
            
            # Additional verification that the cell was written correctly
            verification_value = self.sheet.cell(row=row_index, column=num_col_index).value
            logger.info(f"Cell value verification: Old={old_value}, New={verification_value}, Expected={text}")
            
            # Verify A1 remains unchanged if we're not writing to A1
            if row_index != 1 or num_col_index != 1:
                a1_value = self.sheet.cell(row=1, column=1).value
                logger.info(f"Verifying A1 value remains unchanged: {a1_value}")
            
            success_msg = f"Value '{text}' written to {cell_ref}"
            logger.info(success_msg)
            
            # Save the workbook
            self.workbook.save(self.filename)
            
            return True, success_msg
        except Exception as e:
            error_msg = f"Error writing to cell: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def write_row(self, row_index, row_data):
        """
        Write data to an entire row.
        
        Args:
            row_index (int): Row index (1-based)
            row_data (list): List of values to write to the row
            
        Returns:
            tuple: (success, message)
                - success (bool): True if operation succeeded, False otherwise
                - message (str): Success or error message
        """
        try:
            # Validate row_index
            if not self._validate_row_index(row_index):
                error_msg = f"Invalid row index: {row_index}. Row index must be positive integer."
                logger.error(error_msg)
                return False, error_msg
            
            # Check if row_data is iterable and not a string
            # Strings are iterable but not valid row data
            if isinstance(row_data, str):
                error_msg = f"Row data must be an iterable collection, not a string"
                logger.error(error_msg)
                return False, error_msg
                
            # Check if row_data is iterable
            try:
                iterator = iter(row_data)
            except TypeError:
                error_msg = f"Row data must be iterable, got {type(row_data).__name__}"
                logger.error(error_msg)
                return False, error_msg
            
            # Write data to the row
            for i, value in enumerate(row_data, 1):
                self.sheet.cell(row=row_index, column=i).value = value
            
            # Create a more descriptive message about what was written
            row_data_summary = ", ".join([f"column {get_column_letter(i+1)}: '{val}'" for i, val in enumerate(row_data) if val is not None])
            success_msg = f"Data written to row {row_index}. Values: {row_data_summary}"
            logger.info(success_msg)
            
            # Save the workbook
            self.workbook.save(self.filename)
            
            return True, success_msg
        except Exception as e:
            error_msg = f"Error writing row: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def clear_cell(self, row_index, col_index):
        """
        Clear the content of a specific cell.
        
        Args:
            row_index (int): Row index (1-based)
            col_index (int or str): Column index (1-based) or letter (A, B, etc.)
            
        Returns:
            tuple: (success, message)
                - success (bool): True if operation succeeded, False otherwise
                - message (str): Success or error message
        """
        try:
            # Validate row_index
            if not self._validate_row_index(row_index):
                error_msg = f"Invalid row index: {row_index}. Row index must be positive integer."
                logger.error(error_msg)
                return False, error_msg
            
            # Get the column index if it's a letter
            num_col_index = self._get_col_index(col_index)
            if num_col_index is None:
                error_msg = f"Invalid column index: {col_index}"
                logger.error(error_msg)
                return False, error_msg
            
            # Get column letter for better readability
            col_letter = get_column_letter(num_col_index)
            
            # Create formatted cell reference
            cell_ref = self._format_cell_reference(row_index, col_letter)
            
            # Clear the cell
            self.sheet.cell(row=row_index, column=num_col_index).value = None
            
            success_msg = f"Content cleared from {cell_ref}"
            logger.info(success_msg)
            
            # Save the workbook
            self.workbook.save(self.filename)
            
            return True, success_msg
        except Exception as e:
            error_msg = f"Error clearing cell: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def clear_row(self, row_index):
        """
        Clear/delete an entire row.
        
        Args:
            row_index (int): Row index (1-based)
            
        Returns:
            tuple: (success, message)
                - success (bool): True if operation succeeded, False otherwise
                - message (str): Success or error message
        """
        try:
            # Validate row_index
            if not self._validate_row_index(row_index):
                error_msg = f"Invalid row index: {row_index}. Row index must be positive integer."
                logger.error(error_msg)
                return False, error_msg
            
            # Get actual row index if it's a string digit
            if isinstance(row_index, str) and row_index.isdigit():
                row_index = int(row_index)
            
            # Store the original row contents for better reporting
            original_row, _ = self.read_row(row_index)
            row_data_description = ", ".join([f"'{val}'" for val in original_row if val is not None])
            
            # Delete the row
            self.sheet.delete_rows(row_index)
            
            success_msg = f"Row {row_index} deleted. Original values: {row_data_description}"
            logger.info(success_msg)
            
            # Save the workbook
            self.workbook.save(self.filename)
            
            return True, success_msg
        except Exception as e:
            error_msg = f"Error clearing row: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def clear_column(self, col_index):
        """
        Clear/delete an entire column.
        
        Args:
            col_index (int or str): Column index (1-based) or letter (A, B, etc.)
            
        Returns:
            tuple: (success, message)
                - success (bool): True if operation succeeded, False otherwise
                - message (str): Success or error message
        """
        try:
            # Get the column index if it's a letter
            num_col_index = self._get_col_index(col_index)
            if num_col_index is None:
                error_msg = f"Invalid column index: {col_index}"
                logger.error(error_msg)
                return False, error_msg
            
            # Get column letter for better reporting
            col_letter = get_column_letter(num_col_index)
            
            # Store the original column contents for better reporting
            original_column, _ = self.read_column(col_index)
            column_data_description = ", ".join([f"'{val}'" for val in original_column if val is not None])
            
            # Delete the column
            self.sheet.delete_cols(num_col_index)
            
            success_msg = f"Column {col_letter} (index {num_col_index}) deleted. Original values: {column_data_description}"
            logger.info(success_msg)
            
            # Save the workbook
            self.workbook.save(self.filename)
            
            return True, success_msg
        except Exception as e:
            error_msg = f"Error clearing column: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def read_header_row(self):
        """
        Read the header row (first row) of the sheet.
        
        Returns:
            tuple: (header_row, message)
                - header_row (list): List of header values or None if error
                - message (str): Success or error message
        """
        try:
            # Check if sheet is empty
            if self.sheet.max_row < 1:
                warning_msg = "Sheet is empty, no header row to read"
                logger.warning(warning_msg)
                return [], warning_msg
            
            # Read the header row
            header_row = []
            for cell in self.sheet[1]:
                header_row.append(cell.value)
            
            success_msg = "Header row read successfully"
            logger.info(success_msg)
            
            return header_row, success_msg
        except Exception as e:
            error_msg = f"Error reading header row: {str(e)}"
            logger.error(error_msg)
            return None, error_msg
    
    def read_column(self, col_index):
        """
        Read an entire column.
        
        Args:
            col_index (int or str): Column index (1-based) or letter (A, B, etc.)
            
        Returns:
            tuple: (column_data, message)
                - column_data (list): List of column values or None if error
                - message (str): Success or error message
        """
        try:
            # Get the column index if it's a letter
            num_col_index = self._get_col_index(col_index)
            if num_col_index is None:
                error_msg = f"Invalid column index: {col_index}"
                logger.error(error_msg)
                return None, error_msg
            
            # Check if sheet is empty
            if self.sheet.max_row < 1:
                warning_msg = "Sheet is empty, no column to read"
                logger.warning(warning_msg)
                return [], warning_msg
            
            # Read the column
            column_data = []
            col_letter = get_column_letter(num_col_index)
            for cell in self.sheet[col_letter]:
                column_data.append(cell.value)
            
            success_msg = f"Column {col_index} read successfully"
            logger.info(success_msg)
            
            return column_data, success_msg
        except Exception as e:
            error_msg = f"Error reading column: {str(e)}"
            logger.error(error_msg)
            return None, error_msg
    
    def read_cell(self, row_index, col_index):
        """
        Read the content of a specific cell.
        
        Args:
            row_index (int): Row index (1-based)
            col_index (int or str): Column index (1-based) or letter (A, B, etc.)
            
        Returns:
            tuple: (cell_value, message)
                - cell_value: Value of the cell or None if error
                - message (str): Success or error message
        """
        try:
            # Validate row_index
            if not self._validate_row_index(row_index):
                error_msg = f"Invalid row index: {row_index}. Row index must be positive integer."
                logger.error(error_msg)
                return None, error_msg
            
            # Get the column index if it's a letter
            num_col_index = self._get_col_index(col_index)
            if num_col_index is None:
                error_msg = f"Invalid column index: {col_index}"
                logger.error(error_msg)
                return None, error_msg
            
            # Get column letter for better readability
            col_letter = get_column_letter(num_col_index)
            
            # Create formatted cell reference
            cell_ref = self._format_cell_reference(row_index, col_letter)
            
            # Read the cell
            cell_value = self.sheet.cell(row=row_index, column=num_col_index).value
            
            success_msg = f"Value '{cell_value}' read from {cell_ref}"
            logger.info(success_msg)
            
            return cell_value, success_msg
        except Exception as e:
            error_msg = f"Error reading cell: {str(e)}"
            logger.error(error_msg)
            return None, error_msg
    
    def read_row(self, row_index):
        """
        Read an entire row.
        
        Args:
            row_index (int): Row index (1-based)
            
        Returns:
            tuple: (row_data, message)
                - row_data (list): List of row values or None if error
                - message (str): Success or error message
        """
        try:
            # Validate row_index
            if not self._validate_row_index(row_index):
                error_msg = f"Invalid row index: {row_index}. Row index must be positive integer."
                logger.error(error_msg)
                return None, error_msg
            
            # Get actual row index if it's a string digit
            if isinstance(row_index, str) and row_index.isdigit():
                row_index = int(row_index)
            
            # Check if row exists
            if row_index > self.sheet.max_row:
                warning_msg = f"Row {row_index} does not exist"
                logger.warning(warning_msg)
                return [], warning_msg
            
            # Read the row
            row_data = []
            for cell in self.sheet[row_index]:
                row_data.append(cell.value)
            
            success_msg = f"Row {row_index} read successfully"
            logger.info(success_msg)
            
            return row_data, success_msg
        except Exception as e:
            error_msg = f"Error reading row: {str(e)}"
            logger.error(error_msg)
            return None, error_msg
    
    def get_column_index_by_header(self, header_name):
        """
        Find the column index by header name.
        
        Args:
            header_name (str): Name of the header to find
            
        Returns:
            tuple: (column_index, message)
                - column_index (int): Column index (1-based) or None if not found
                - message (str): Success or error message
        """
        try:
            # Read the header row
            header_row, _ = self.read_header_row()
            
            # Check if header_row is empty
            if not header_row:
                error_msg = "No header row found"
                logger.error(error_msg)
                return None, error_msg
            
            # Find the header
            for i, header in enumerate(header_row, 1):
                if header == header_name:
                    success_msg = f"Column index found by header: {header_name}"
                    logger.info(success_msg)
                    return i, success_msg
            
            error_msg = f"Header '{header_name}' not found"
            logger.error(error_msg)
            return None, error_msg
        except Exception as e:
            error_msg = f"Error finding column index: {str(e)}"
            logger.error(error_msg)
            return None, error_msg
    
    def get_row_index_by_value(self, col_index, search_value):
        """
        Find a row index by searching for a value in a specific column.
        
        Args:
            col_index (int or str): Column index or letter where to search
            search_value (str): Value to search for
            
        Returns:
            tuple: (row_index, message)
                - row_index (int): Row index (1-based) or None if not found
                - message (str): Success or error message
        """
        try:
            # Convert column index if needed
            numeric_col_index = self._get_col_index(col_index)
            if not numeric_col_index:
                error_msg = f"Invalid column index: {col_index}"
                logger.error(error_msg)
                return None, error_msg
            
            # Read the column
            column_data, _ = self.read_column(col_index)
            if not column_data:
                error_msg = f"No data found in column {col_index}"
                logger.error(error_msg)
                return None, error_msg
            
            # Convert search_value to string for comparison
            search_value_str = str(search_value)
            
            # Find the row with matching value
            for row_idx, cell_value in enumerate(column_data, 1):
                if cell_value is not None and str(cell_value) == search_value_str:
                    success_msg = f"Row index found: {row_idx} with value '{search_value}' in column {col_index}"
                    logger.info(success_msg)
                    return row_idx, success_msg
            
            error_msg = f"Value '{search_value}' not found in column {col_index}"
            logger.error(error_msg)
            return None, error_msg
        except Exception as e:
            error_msg = f"Error finding row index: {str(e)}"
            logger.error(error_msg)
            return None, error_msg
    
    def update_cell_by_lookup(self, row_header, row_value, col_header, new_value):
        """
        Find and update a cell by looking up the row based on a value and column based on header.
        
        Args:
            row_header (str): Header name of the column that contains the row identifier
            row_value (str): Value to search for in the row identifier column
            col_header (str): Header name of the column to update
            new_value (str): New value to set in the target cell
            
        Returns:
            tuple: (success, message)
                - success (bool): True if the operation was successful, False otherwise
                - message (str): Success or error message with details of the operation
        """
        try:
            # Step 1: Find column index for row identifier
            row_col_idx, message = self.get_column_index_by_header(row_header)
            if not row_col_idx:
                error_msg = f"Could not find column with header '{row_header}': {message}"
                logger.error(error_msg)
                return False, error_msg
            
            # Step 2: Find column index for target column
            target_col_idx, message = self.get_column_index_by_header(col_header)
            if not target_col_idx:
                error_msg = f"Could not find column with header '{col_header}': {message}"
                logger.error(error_msg)
                return False, error_msg
            
            # Step 3: Find row index based on value
            row_idx, message = self.get_row_index_by_value(row_col_idx, row_value)
            if not row_idx:
                error_msg = f"Could not find row with value '{row_value}' in column '{row_header}': {message}"
                logger.error(error_msg)
                return False, error_msg
            
            # Step 4: Update the target cell
            success, write_message = self.write_cell(row_idx, target_col_idx, new_value)
            if not success:
                error_msg = f"Failed to write to cell at row {row_idx}, column {target_col_idx}: {write_message}"
                logger.error(error_msg)
                return False, error_msg
            
            # Create a human-readable message with cell coordinates
            col_letter = get_column_letter(target_col_idx)
            success_msg = (f"Successfully updated cell at row {row_idx}, column {col_letter} "
                          f"(identified by '{row_header}={row_value}' and '{col_header}') "
                          f"with value '{new_value}'")
            logger.info(success_msg)
            return True, success_msg
            
        except Exception as e:
            error_msg = f"Error updating cell by lookup: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    #
    # JSON OPERATION PROCESSING
    #
    
    def process_json_operation(self, json_input):
        """
        Process a JSON-formatted Excel operation and return a reward.
        
        The JSON format should be:
        {
            "function_name": "excel_function_name",
            "parameters": {
                "param1": value1,
                "param2": value2,
                ...
            }
        }
        
        Args:
            json_input (str): JSON-formatted operation
            
        Returns:
            tuple: (reward, feedback)
                - reward (int): 1 for success, -1 for failure
                - feedback (str): Detailed feedback about the operation
        """
        try:
            # Parse the JSON
            try:
                operation = json.loads(json_input)
                logger.info(f"Processing JSON operation: {json_input}")
            except json.JSONDecodeError:
                error_msg = "Invalid JSON format"
                logger.error(error_msg)
                return -1, f"Error: {error_msg}"
            
            # Check if function_name is present
            if "function_name" not in operation:
                error_msg = "JSON missing 'function_name' field"
                logger.error(error_msg)
                return -1, f"Error: {error_msg}"
            
            function_name = operation["function_name"]
            parameters = operation.get("parameters", {})
            
            # Log parameter validation
            logger.info(f"Function: {function_name}, Parameters: {parameters}")
            
            # Process the operation based on function_name
            if function_name == "excel_clear_sheet":
                success, message = self.clear_sheet()
            
            elif function_name == "excel_add_row":
                # Check required parameters
                if not self._validate_parameters(parameters, ["row_index", "text"]):
                    error_msg = "Missing required parameters for add_row. Needs: row_index, text"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                # Extra validation for row_index
                row_index = parameters["row_index"]
                if not (row_index == "next_available" or 
                        (isinstance(row_index, int) and row_index > 0) or
                        (isinstance(row_index, str) and row_index.isdigit() and int(row_index) > 0)):
                    error_msg = f"Invalid row_index: {row_index}. Must be positive integer or 'next_available'"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                success, message = self.add_row(
                    parameters["row_index"],
                    parameters["text"]
                )
            
            elif function_name == "excel_write_cell":
                # Check required parameters
                if not self._validate_parameters(parameters, ["row_index", "col_index", "text"]):
                    error_msg = "Missing required parameters for write_cell. Needs: row_index, col_index, text"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                # Extract parameters with detailed logging
                row_index = parameters["row_index"]
                col_index = parameters["col_index"]
                text = parameters["text"]
                
                logger.info(f"JSON WRITE_CELL PARAMETERS - row_index: {row_index} ({type(row_index).__name__}), " +
                            f"col_index: {col_index} ({type(col_index).__name__}), text: {text}")
                
                # Extra validation for row_index
                if not (isinstance(row_index, int) and row_index > 0 or 
                        isinstance(row_index, str) and row_index.isdigit() and int(row_index) > 0):
                    error_msg = f"Invalid row_index: {row_index}. Must be positive integer"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                # Extra validation for col_index
                if isinstance(col_index, str):
                    # If it's a letter, make sure it's a valid column letter
                    if not col_index.isdigit() and (len(col_index) > 3 or not all(c.isalpha() for c in col_index)):
                        error_msg = f"Invalid col_index: {col_index}. Must be a column letter (A-Z) or positive integer"
                        logger.error(error_msg)
                        return -1, f"Error: {error_msg}"
                elif not (isinstance(col_index, int) and col_index > 0):
                    error_msg = f"Invalid col_index: {col_index}. Must be positive integer or column letter"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                # Convert row_index to integer if it's a string digit
                if isinstance(row_index, str) and row_index.isdigit():
                    row_index = int(row_index)
                    parameters["row_index"] = row_index
                    logger.info(f"Converted row_index string to int: {row_index}")
                
                # Log the exact cell we're targeting
                logger.info(f"JSON WRITE TARGETING: Row {row_index}, Column {col_index}")
                
                # Call the write_cell method with explicit parameter names
                success, message = self.write_cell(
                    row_index=row_index,
                    col_index=col_index,
                    text=text
                )
            
            elif function_name == "excel_write_row":
                # Check required parameters
                if not self._validate_parameters(parameters, ["row_index", "row_data"]):
                    error_msg = "Missing required parameters for write_row. Needs: row_index, row_data"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                # Validate row_data is iterable
                try:
                    iter(parameters["row_data"])
                except TypeError:
                    error_msg = f"Invalid row_data: {parameters['row_data']}. Must be iterable (list, tuple, etc.)"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                success, message = self.write_row(
                    parameters["row_index"],
                    parameters["row_data"]
                )
            
            elif function_name == "excel_clear_cell":
                # Check required parameters
                if not self._validate_parameters(parameters, ["row_index", "col_index"]):
                    error_msg = "Missing required parameters for clear_cell. Needs: row_index, col_index"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                success, message = self.clear_cell(
                    parameters["row_index"],
                    parameters["col_index"]
                )
            
            elif function_name == "excel_clear_row":
                # Check required parameters
                if not self._validate_parameters(parameters, ["row_index"]):
                    error_msg = "Missing required parameters for clear_row. Needs: row_index"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                success, message = self.clear_row(
                    parameters["row_index"]
                )
            
            elif function_name == "excel_clear_column":
                # Check required parameters
                if not self._validate_parameters(parameters, ["col_index"]):
                    error_msg = "Missing required parameters for clear_column. Needs: col_index"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                success, message = self.clear_column(
                    parameters["col_index"]
                )
            
            elif function_name == "excel_read_header_row":
                result, message = self.read_header_row()
                success = result is not None
                
                # Format result for feedback
                if success:
                    # Create a more descriptive message about the header contents
                    header_description = ", ".join([f"'{h}'" for h in result])
                    message = f"Success: Header row read successfully. Headers found: {header_description}"
            
            elif function_name == "excel_read_column":
                # Check required parameters
                if not self._validate_parameters(parameters, ["col_index"]):
                    error_msg = "Missing required parameters for read_column. Needs: col_index"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                col_index = parameters["col_index"]
                
                # Get a readable column identifier for the message
                if isinstance(col_index, int):
                    col_id = f"column {col_index} ({get_column_letter(col_index)})"
                elif isinstance(col_index, str) and col_index.isalpha():
                    col_id = f"column {col_index}"
                else:
                    col_id = f"column {col_index}"
                
                result, message = self.read_column(
                    parameters["col_index"]
                )
                success = result is not None
                
                # Format result for feedback
                if success:
                    # Format column data in a more descriptive way
                    column_data_summary = ", ".join([f"row {i+1}: '{val}'" for i, val in enumerate(result) if val is not None])
                    message = f"Success: {col_id} read successfully. Values: {column_data_summary}"
            
            elif function_name == "excel_read_cell":
                # Check required parameters
                if not self._validate_parameters(parameters, ["row_index", "col_index"]):
                    error_msg = "Missing required parameters for read_cell. Needs: row_index, col_index"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                row_index = parameters["row_index"]
                col_index = parameters["col_index"]
                
                result, message = self.read_cell(
                    row_index,
                    col_index
                )
                success = result is not None or message.startswith("Value")
                
                # Format result for feedback - use cell_ref format
                if success:
                    # Get column letter for better formatting
                    num_col_index = self._get_col_index(col_index)
                    if num_col_index:
                        col_letter = get_column_letter(num_col_index)
                        cell_ref = self._format_cell_reference(row_index, col_letter)
                        message = f"Success: Read value '{result}' from {cell_ref}"
                    else:
                        message = f"Success: Cell read successfully. Result: {result}"
            
            elif function_name == "excel_read_row":
                # Check required parameters
                if not self._validate_parameters(parameters, ["row_index"]):
                    error_msg = "Missing required parameters for read_row. Needs: row_index"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                row_index = parameters["row_index"]
                
                result, message = self.read_row(
                    row_index
                )
                success = result is not None
                
                # Format result for feedback
                if success:
                    # Format row data in a more descriptive way
                    row_data_summary = ", ".join([f"column {get_column_letter(i+1)}: '{val}'" for i, val in enumerate(result) if val is not None])
                    message = f"Success: Row {row_index} read successfully. Values: {row_data_summary}"
            
            elif function_name == "excel_get_column_index_by_header":
                # Check required parameters
                if not self._validate_parameters(parameters, ["header_name"]):
                    error_msg = "Missing required parameters for get_column_index_by_header. Needs: header_name"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                result, message = self.get_column_index_by_header(
                    parameters["header_name"]
                )
                success = result is not None
                
                # Format result for feedback
                if success:
                    message = f"Success: Column index found by header. Result: {result}"
            
            elif function_name == "excel_get_row_index_by_value":
                # Check required parameters
                if not self._validate_parameters(parameters, ["col_index", "search_value"]):
                    error_msg = "Missing required parameters for get_row_index_by_value. Needs: col_index, search_value"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                col_index = parameters["col_index"]
                search_value = parameters["search_value"]
                
                result, message = self.get_row_index_by_value(
                    col_index,
                    search_value
                )
                success = result is not None
                
                # Format result for feedback
                if success:
                    message = f"Success: Row index found by value. Result: {result}"
            
            elif function_name == "excel_update_cell_by_lookup":
                # Check required parameters
                if not self._validate_parameters(parameters, ["row_header", "row_value", "col_header", "new_value"]):
                    error_msg = "Missing required parameters for update_cell_by_lookup. Needs: row_header, row_value, col_header, new_value"
                    logger.error(error_msg)
                    return -1, f"Error: {error_msg}"
                
                row_header = parameters["row_header"]
                row_value = parameters["row_value"]
                col_header = parameters["col_header"]
                new_value = parameters["new_value"]
                
                result, message = self.update_cell_by_lookup(
                    row_header,
                    row_value,
                    col_header,
                    new_value
                )
                success = result
                
                # Format result for feedback
                if success:
                    message = f"Success: Cell updated successfully. {message}"
            
            else:
                error_msg = f"Unknown function: {function_name}"
                logger.error(error_msg)
                return -1, f"Error: {error_msg}"
            
            # Calculate reward based on success
            reward = 1 if success else -1
            
            # Format feedback
            if success:
                if not message.startswith("Success"):
                    feedback = f"Success: {message}"
                else:
                    feedback = message
            else:
                feedback = f"Error: {message}"
            
            # Log the final result
            logger.info(f"Operation result: reward={reward}, feedback={feedback}")
            
            return reward, feedback
            
        except Exception as e:
            error_msg = f"Error processing JSON operation: {str(e)}"
            logger.error(error_msg)
            return -1, f"Error: {error_msg}"